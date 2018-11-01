"""
pip install pyexcel-xlsx
pyexcel-io==0.5.9.1
pyexcel-xlsx==0.5.6
python professional-review.py
Packages need to install to run webdriver
sudo apt-get install python3-pip
pip3 install selenium
sudo apt-get install firefox
sudo apt-get install xvfb
sudo pip install pyvirtualdisplay
TO ENTER AND COME BACK FROM screen on server
1) To resume screen $ screen -r or for new $ screen
2) To come back CTR+A then press d

"""
import re
import os
import time
import requests
from bs4 import BeautifulSoup
from pyvirtualdisplay import Display
from selenium import webdriver
import unicodedata
import json
import logging
from pyexcel_xlsx import save_data
from collections import OrderedDict
logging.basicConfig(filename="job-logs.log",
                    format='%(asctime)s %(levelname)s %(message)s',
                    filemode='w',
                    level=logging.INFO)
logger=logging.getLogger()

BASE_URL = "http://professional-review.com/"
COUNTRY_LIST=[]
HEADER_LIST = ["Country", "State", "Store Name", "Address", "Website URL",
"Tel Phone", "Email", "About", "Description"]
data = OrderedDict()


display = Display(visible=0, size=(1024, 768))
display.start()
capabilities = webdriver.DesiredCapabilities().FIREFOX
capabilities["marionette"] = True
capabilities["binary"] = '/usr/bin/firefox'
browser = webdriver.Firefox(capabilities= capabilities)

def get_country():
	logger.info("#----start to find all country----") 
	req = requests.get(BASE_URL)
	if req.status_code == 200:
		response = req.content;
		soup = BeautifulSoup(response, 'html.parser');
		links = soup.find_all("a", class_="p-2 text-muted")
		for link in links:
			COUNTRY_LIST.append({"country":str(link['href'].split("/")[0].title()), "state": str(link.text),"link":BASE_URL+link['href']})
		logger.info("#----country find success----")
	else:
		logger.error("#----country find request failed.----")

def get_details():
	get_country()
	for country in COUNTRY_LIST:
		req = requests.get(country['link'])
		if req.status_code == 200:
			UPDATE_DATA = [HEADER_LIST]
			response = req.content;
			soup = BeautifulSoup(response, 'html.parser');
			
			updates = soup.find_all("div", class_="card flex-md-row mb-4 box-shadow h-md-200")
			for update in updates:
				header_link = BASE_URL+update.find('strong').find("a")['href']
				browser.get(header_link)
				time.sleep(5)
				update_response = browser.page_source
				update_soup = BeautifulSoup(update_response, 'html.parser');
				header_sec = update_soup.find("div", class_="jumbotron")
				store_name = header_sec.find("h1", class_="display-4").text.strip()
				store_name = unicodedata.normalize('NFKD', store_name).encode('ascii','ignore')
				address = header_sec.find("p", class_="lead my-3").text.strip()
				address = unicodedata.normalize('NFKD', address).encode('ascii','ignore')
				other_tele = header_sec.find_all("a")
				web_url = ".."
				tel_phone = ".."
				email = ".."
				for i in range(len(other_tele)):
					if other_tele[i]['href'].startswith('http'):
						web_url = other_tele[i]['href'].strip()
						web_url = unicodedata.normalize('NFKD', web_url).encode('ascii','ignore')

					if  other_tele[i]['href'].startswith('tel:'):
						tel_phone = other_tele[i]['href'].split(":")[1].strip()
						tel_phone = unicodedata.normalize('NFKD', tel_phone).encode('ascii','ignore')

					if "@" in other_tele[i]['href']:
						email = other_tele[i].text.strip()
						email = unicodedata.normalize('NFKD', email).encode('ascii','ignore')

				about = ".."
				description = ".." 
				about_desc = update_soup.select("div.p-3.mb-3.bg-light.rounded")
				for i in range(len(about_desc)):
					if about_desc[i].find("h4").text.strip() == 'Description':
						description = about_desc[i].find("p").text.strip().replace('\n', ' ')
						description = unicodedata.normalize('NFKD', description).encode('ascii','ignore')

					if about_desc[i].find("h4").text.strip() == 'About':
						about = about_desc[i].find("p").text.strip().replace('\n', ' ')
						about = unicodedata.normalize('NFKD', about).encode('ascii','ignore')
				
				UPDATE_DATA.append([country['country'],country['state']
				, store_name, address, web_url,
				tel_phone, email, about, description])
			try:
				FILE_NAME = os.path.join(os.getcwd(),"report/professional-review-{}.xlsx".format(country['state']))

				save_data(FILE_NAME, {"Sheet {}".format(country['state']): UPDATE_DATA})
				format_file(FILE_NAME)
			except Exception as ex:
				logger.error("#---save error---{}".format(ex))
		else:
			logger.error("#----country {0} find details failed link {1}.----".format(country['country'], country['link']))


def format_file(FileName):
	from openpyxl import load_workbook
	from openpyxl.styles import Font, Alignment
	wb = load_workbook(filename=FileName)
	black_font = Font(size=11, bold=True, color='FF000000')
	count = 1
	
	for ws in wb.worksheets:
		width = 15
		ws.row_dimensions[0].height = 30
		for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
			ws.column_dimensions[column].width = width
			width+=10

		for cell in ws["1:1"]:
			cell.font = black_font
			cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
			count +=1

	wb.save(filename=FileName)

if __name__ == "__main__":
	get_details()
	logger.info("#-------done--------")
	display.stop()
	browser.quit()